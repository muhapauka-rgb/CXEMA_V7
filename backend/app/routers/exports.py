from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from ..db import get_db
from ..models import (
    ClientBillingAdjustment,
    ClientPaymentsFact,
    ClientPaymentsPlan,
    ExpenseGroup,
    ExpenseItem,
    ItemMode,
    Project,
)
from ..utils import compute_project_financials

router = APIRouter(prefix="/api/exports", tags=["exports"])


def _item_base_total(item: ExpenseItem) -> float:
    base = float(item.base_total or 0.0)
    if item.mode == ItemMode.QTY_PRICE and item.unit_price_base is not None:
        qty = float(item.qty or 0.0)
        unit = float(item.unit_price_base or 0.0)
        return unit if qty == 0 else qty * unit
    return base


def _month_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        return f"{value.year:04d}-{value.month:02d}"
    except Exception:
        return ""


def _safe_num(value: Any) -> float:
    try:
        return float(value or 0.0)
    except Exception:
        return 0.0


def _str_bool(flag: Any) -> str:
    return "Да" if bool(flag) else "Нет"


@router.get("/excel")
def export_full_registry_excel(db: Session = Depends(get_db)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except Exception as exc:  # pragma: no cover
        raise HTTPException(status_code=503, detail="OPENPYXL_REQUIRED") from exc

    projects = db.execute(select(Project).order_by(Project.id.asc())).scalars().all()
    groups = db.execute(select(ExpenseGroup).order_by(ExpenseGroup.project_id.asc(), ExpenseGroup.sort_order.asc(), ExpenseGroup.id.asc())).scalars().all()
    items = db.execute(select(ExpenseItem).order_by(ExpenseItem.project_id.asc(), ExpenseItem.group_id.asc(), ExpenseItem.id.asc())).scalars().all()
    adjustments = db.execute(select(ClientBillingAdjustment)).scalars().all()
    payments_fact = db.execute(select(ClientPaymentsFact).order_by(ClientPaymentsFact.project_id.asc(), ClientPaymentsFact.pay_date.asc(), ClientPaymentsFact.id.asc())).scalars().all()
    payments_plan = db.execute(select(ClientPaymentsPlan).order_by(ClientPaymentsPlan.project_id.asc(), ClientPaymentsPlan.pay_date.asc(), ClientPaymentsPlan.id.asc())).scalars().all()

    project_by_id = {int(p.id): p for p in projects}
    group_by_id = {int(g.id): g for g in groups}
    item_by_id = {int(it.id): it for it in items}
    adjustment_by_item_id = {int(adj.expense_item_id): adj for adj in adjustments}

    wb = Workbook()
    ws_ops = wb.active
    ws_ops.title = "Операции"

    ops_headers = [
        "Дата",
        "Период",
        "Проект",
        "Организация",
        "Категория",
        "Источник",
        "Группа",
        "Статья",
        "Родительская статья",
        "Шт",
        "Цена за ед",
        "База",
        "Доп прибыль",
        "Скидка",
        "Итог строки",
        "Влияние на баланс",
        "В смету",
        "Примечание",
    ]
    ws_ops.append(ops_headers)

    ops_rows: list[list[Any]] = []

    for rec in payments_fact:
        project = project_by_id.get(int(rec.project_id))
        if not project:
            continue
        amount = _safe_num(rec.amount)
        ops_rows.append(
            [
                rec.pay_date,
                _month_text(rec.pay_date),
                project.title,
                project.client_name or "—",
                "Приход",
                "Оплата факт",
                "",
                "",
                "",
                "",
                "",
                amount,
                "",
                "",
                amount,
                amount,
                "",
                rec.note or "",
            ]
        )

    for rec in payments_plan:
        project = project_by_id.get(int(rec.project_id))
        if not project:
            continue
        amount = _safe_num(rec.amount)
        ops_rows.append(
            [
                rec.pay_date,
                _month_text(rec.pay_date),
                project.title,
                project.client_name or "—",
                "Приход",
                "Оплата план",
                "",
                "",
                "",
                "",
                "",
                amount,
                "",
                "",
                amount,
                amount,
                "",
                rec.note or "",
            ]
        )

    for item in items:
        project = project_by_id.get(int(item.project_id))
        group = group_by_id.get(int(item.group_id))
        if not project or not group:
            continue

        parent_title = ""
        if item.parent_item_id is not None:
            parent = item_by_id.get(int(item.parent_item_id))
            if parent:
                parent_title = parent.title

        adj = adjustment_by_item_id.get(int(item.id))
        discount_enabled = bool(adj.discount_enabled) if adj else False
        discount_amount = _safe_num(adj.discount_amount) if discount_enabled else 0.0
        base = _item_base_total(item)
        extra = _safe_num(item.extra_profit_amount) if bool(item.extra_profit_enabled) else 0.0
        row_total = base + extra - discount_amount

        ops_rows.append(
            [
                item.planned_pay_date,
                _month_text(item.planned_pay_date),
                project.title,
                project.client_name or "—",
                "Расход",
                "Позиция",
                group.name,
                item.title,
                parent_title,
                "" if item.qty is None else _safe_num(item.qty),
                "" if item.unit_price_base is None else _safe_num(item.unit_price_base),
                base,
                extra if extra != 0 else "",
                discount_amount if discount_enabled else "",
                row_total,
                -row_total,
                _str_bool(item.include_in_estimate),
                "",
            ]
        )

    ops_rows.sort(
        key=lambda row: (
            str(row[0] or "9999-99-99"),
            str(row[2] or ""),
            str(row[5] or ""),
            str(row[7] or ""),
        )
    )

    for row in ops_rows:
        ws_ops.append(row)

    ws_projects = wb.create_sheet("Свод по проектам")
    project_headers = [
        "Проект",
        "Организация",
        "Стоимость проекта",
        "Получено (факт+план)",
        "Потрачено (с УСН)",
        "Агентские",
        "Доп прибыль",
        "Скидка",
        "УСН",
        "В кармане",
        "Баланс",
        "Дата закрытия",
    ]
    ws_projects.append(project_headers)

    payment_fact_sum = {
        int(pid): float(total or 0.0)
        for pid, total in db.execute(
            select(ClientPaymentsFact.project_id, func.coalesce(func.sum(ClientPaymentsFact.amount), 0.0))
            .group_by(ClientPaymentsFact.project_id)
        ).all()
    }
    payment_plan_sum = {
        int(pid): float(total or 0.0)
        for pid, total in db.execute(
            select(ClientPaymentsPlan.project_id, func.coalesce(func.sum(ClientPaymentsPlan.amount), 0.0))
            .group_by(ClientPaymentsPlan.project_id)
        ).all()
    }

    org_agg: dict[str, dict[str, float]] = defaultdict(lambda: {
        "projects": 0.0,
        "project_total": 0.0,
        "received_total": 0.0,
        "expenses_total": 0.0,
        "agency_fee": 0.0,
        "extra_profit_total": 0.0,
        "discount_total": 0.0,
        "usn_tax": 0.0,
        "in_pocket": 0.0,
        "diff": 0.0,
    })

    for project in projects:
        fin = compute_project_financials(db, int(project.id))
        received = float(payment_fact_sum.get(int(project.id), 0.0)) + float(payment_plan_sum.get(int(project.id), 0.0))
        row = [
            project.title,
            project.client_name or "—",
            _safe_num(project.project_price_total),
            received,
            _safe_num(fin.get("expenses_total")),
            _safe_num(fin.get("agency_fee")),
            _safe_num(fin.get("extra_profit_total")),
            _safe_num(fin.get("discount_total")),
            _safe_num(fin.get("usn_tax")),
            _safe_num(fin.get("in_pocket")),
            _safe_num(fin.get("diff")),
            project.closed_at,
        ]
        ws_projects.append(row)

        org_name = project.client_name or "—"
        agg = org_agg[org_name]
        agg["projects"] += 1
        agg["project_total"] += row[2]
        agg["received_total"] += row[3]
        agg["expenses_total"] += row[4]
        agg["agency_fee"] += row[5]
        agg["extra_profit_total"] += row[6]
        agg["discount_total"] += row[7]
        agg["usn_tax"] += row[8]
        agg["in_pocket"] += row[9]
        agg["diff"] += row[10]

    ws_org = wb.create_sheet("Свод по фирмам")
    org_headers = [
        "Организация",
        "Проектов",
        "Стоимость проектов",
        "Получено (факт+план)",
        "Потрачено (с УСН)",
        "Агентские",
        "Доп прибыль",
        "Скидка",
        "УСН",
        "В кармане",
        "Баланс",
    ]
    ws_org.append(org_headers)

    for org_name in sorted(org_agg.keys(), key=lambda x: x.lower()):
        agg = org_agg[org_name]
        ws_org.append(
            [
                org_name,
                int(agg["projects"]),
                agg["project_total"],
                agg["received_total"],
                agg["expenses_total"],
                agg["agency_fee"],
                agg["extra_profit_total"],
                agg["discount_total"],
                agg["usn_tax"],
                agg["in_pocket"],
                agg["diff"],
            ]
        )

    header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_side = Side(style="thin", color="D0D7DE")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def style_sheet(
        ws: Any,
        *,
        money_cols: set[int] | None = None,
        date_cols: set[int] | None = None,
        center_cols: set[int] | None = None,
        width_map: dict[int, float] | None = None,
    ) -> None:
        money_cols = money_cols or set()
        date_cols = date_cols or set()
        center_cols = center_cols or set()
        width_map = width_map or {}

        max_row = ws.max_row
        max_col = ws.max_column
        if max_row < 1 or max_col < 1:
            return

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                col = cell.column
                cell.border = border
                if col in date_cols:
                    cell.number_format = "yyyy-mm-dd"
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col in money_cols:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col in center_cols:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        for idx in range(1, max_col + 1):
            width = width_map.get(idx)
            if width is not None:
                ws.column_dimensions[get_column_letter(idx)].width = width

    style_sheet(
        ws_ops,
        money_cols={10, 11, 12, 13, 14, 15, 16},
        date_cols={1},
        center_cols={17},
        width_map={
            1: 12, 2: 10, 3: 22, 4: 22, 5: 11, 6: 14, 7: 16, 8: 22, 9: 22,
            10: 10, 11: 12, 12: 12, 13: 12, 14: 12, 15: 12, 16: 14, 17: 10, 18: 24,
        },
    )
    style_sheet(
        ws_projects,
        money_cols={3, 4, 5, 6, 7, 8, 9, 10, 11},
        date_cols={12},
        width_map={1: 24, 2: 20, 3: 15, 4: 18, 5: 18, 6: 12, 7: 12, 8: 12, 9: 12, 10: 12, 11: 12, 12: 14},
    )
    style_sheet(
        ws_org,
        money_cols={3, 4, 5, 6, 7, 8, 9, 10, 11},
        center_cols={2},
        width_map={1: 26, 2: 10, 3: 18, 4: 18, 5: 18, 6: 12, 7: 12, 8: 12, 9: 12, 10: 12, 11: 12},
    )

    stream = BytesIO()
    wb.save(stream)
    content = stream.getvalue()
    stamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    filename = f"cxema-registry-{stamp}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
