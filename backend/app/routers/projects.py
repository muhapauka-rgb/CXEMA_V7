from __future__ import annotations

import csv
import io
import json
import re
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from sqlalchemy.orm import Session
from sqlalchemy import delete, select, text, func, update
from sqlalchemy.exc import SQLAlchemyError

from ..db import get_db, engine, Base
from ..models import (
    Project,
    ExpenseGroup,
    ExpenseItem,
    ClientBillingAdjustment,
    ClientPaymentsPlan,
    ClientPaymentsFact,
    ItemMode,
    AdjustmentType,
)
from ..schemas import (
    ProjectCreate,
    ProjectUpdate,
    ProjectOut,
    GroupCreate,
    GroupUpdate,
    GroupOut,
    ItemCreate,
    ItemUpdate,
    ItemOut,
    BillingAdjustmentUpsert,
    BillingAdjustmentOut,
    PaymentPlanCreate,
    PaymentPlanUpdate,
    PaymentPlanOut,
    PaymentFactCreate,
    PaymentFactUpdate,
    PaymentFactOut,
    ProjectComputed,
    ProjectReorderIn,
    ContractorEstimateImportOut,
    ContractorEstimatePreviewOut,
)
from ..utils import compute_project_financials, gen_stable_id

router = APIRouter(prefix="/api/projects", tags=["projects"])

# MVP: create tables automatically on first import (без alembic пока)
Base.metadata.create_all(bind=engine)


def _ensure_sqlite_columns() -> None:
    if engine.dialect.name != "sqlite":
        return
    # Lightweight compatibility migration for existing SQLite files.
    with engine.begin() as conn:
        project_columns = {row[1] for row in conn.execute(text("PRAGMA table_info(projects)"))}
        if "client_email" not in project_columns:
            conn.execute(text("ALTER TABLE projects ADD COLUMN client_email VARCHAR(255)"))
        if "client_phone" not in project_columns:
            conn.execute(text("ALTER TABLE projects ADD COLUMN client_phone VARCHAR(64)"))
        if "google_drive_url" not in project_columns:
            conn.execute(text("ALTER TABLE projects ADD COLUMN google_drive_url VARCHAR(1024)"))
        if "google_drive_folder" not in project_columns:
            conn.execute(text("ALTER TABLE projects ADD COLUMN google_drive_folder VARCHAR(255)"))
        if "card_image_data" not in project_columns:
            conn.execute(text("ALTER TABLE projects ADD COLUMN card_image_data TEXT"))
        if "agency_fee_percent" not in project_columns:
            conn.execute(text("ALTER TABLE projects ADD COLUMN agency_fee_percent FLOAT NOT NULL DEFAULT 10.0"))
        if "agency_fee_include_in_estimate" not in project_columns:
            conn.execute(text("ALTER TABLE projects ADD COLUMN agency_fee_include_in_estimate BOOLEAN NOT NULL DEFAULT 1"))
        if "sort_order" not in project_columns:
            conn.execute(text("ALTER TABLE projects ADD COLUMN sort_order INTEGER NOT NULL DEFAULT 0"))
        conn.execute(text("UPDATE projects SET sort_order = id WHERE sort_order IS NULL OR sort_order = 0"))
        if "is_paused" not in project_columns:
            conn.execute(text("ALTER TABLE projects ADD COLUMN is_paused BOOLEAN NOT NULL DEFAULT 0"))

        columns = {row[1] for row in conn.execute(text("PRAGMA table_info(expense_items)"))}
        if "include_in_estimate" not in columns:
            conn.execute(
                text(
                    "ALTER TABLE expense_items "
                    "ADD COLUMN include_in_estimate BOOLEAN NOT NULL DEFAULT 1"
                )
            )
        if "parent_item_id" not in columns:
            conn.execute(
                text(
                    "ALTER TABLE expense_items "
                    "ADD COLUMN parent_item_id INTEGER"
                )
            )
        adjustment_columns = {row[1] for row in conn.execute(text("PRAGMA table_info(client_billing_adjustments)"))}
        if "discount_enabled" not in adjustment_columns:
            conn.execute(
                text(
                    "ALTER TABLE client_billing_adjustments "
                    "ADD COLUMN discount_enabled BOOLEAN NOT NULL DEFAULT 0"
                )
            )
        if "discount_amount" not in adjustment_columns:
            conn.execute(
                text(
                    "ALTER TABLE client_billing_adjustments "
                    "ADD COLUMN discount_amount FLOAT NOT NULL DEFAULT 0.0"
                )
            )


_ensure_sqlite_columns()

_ROW_HEADER_HINTS = {
    "статья",
    "наименование",
    "название",
    "позиция",
    "кол-во",
    "количество",
    "ед",
    "ед.",
    "цена",
    "сумма",
    "итого",
}

_AGGREGATE_ROW_HINTS = (
    "итого",
    "всего",
    "сумма",
    "subtotal",
    "total",
    "ндс",
    "налог",
    "налоги",
)


def _norm_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _parse_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = _norm_text(value)
    if not raw:
        return None
    # Ignore mixed text like "2 ой этаж" so digits inside labels are not parsed as amounts.
    normalized = raw.replace("\u00A0", "").replace(" ", "").replace(",", ".")
    if re.search(r"[A-Za-zА-Яа-я]", normalized):
        return None
    normalized = re.sub(r"[^0-9.\-]", "", normalized)
    if not re.fullmatch(r"-?\d+(\.\d+)?", normalized):
        return None
    if not normalized:
        return None
    try:
        return float(normalized)
    except Exception:
        return None


def _is_aggregate_or_tax_row_label(label: str) -> bool:
    text = _norm_text(label).lower()
    if not text:
        return False
    return any(token in text for token in _AGGREGATE_ROW_HINTS)


def _load_table_rows_from_file(filename: str, content: bytes) -> list[list[object]]:
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if ext in {"xlsx", "xlsm", "xltx", "xltm"}:
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise HTTPException(status_code=500, detail="OPENPYXL_NOT_INSTALLED") from exc
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = wb.active
        out: list[list[object]] = []
        for row in ws.iter_rows(values_only=True):
            out.append(list(row))
        return out

    if ext in {"csv", "tsv", "txt"}:
        text: str | None = None
        for enc in ("utf-8-sig", "cp1251", "utf-8"):
            try:
                text = content.decode(enc)
                break
            except Exception:
                continue
        if text is None:
            text = content.decode("latin-1", errors="ignore")
        sample = "\n".join(text.splitlines()[:5])
        delimiters = {";": sample.count(";"), ",": sample.count(","), "\t": sample.count("\t")}
        delim = max(delimiters.items(), key=lambda kv: kv[1])[0] if delimiters else ";"
        reader = csv.reader(io.StringIO(text), delimiter=delim)
        return [list(r) for r in reader]

    raise HTTPException(status_code=422, detail="UNSUPPORTED_ESTIMATE_FILE_FORMAT")


def _parse_imported_estimate_rows(rows: list[list[object]], fallback_block_name: str) -> list[dict]:
    blocks: list[dict] = []
    current_block: dict | None = None

    def ensure_block(name: str) -> dict:
        nonlocal current_block
        title = _norm_text(name) or fallback_block_name
        if current_block and current_block["title"] == title:
            return current_block
        block = {"title": title, "items": [], "explicit_total": None}
        blocks.append(block)
        current_block = block
        return block

    for row in rows:
        values = [v for v in row if _norm_text(v)]
        if not values:
            continue

        text_cells = [_norm_text(v) for v in row if isinstance(v, str) and _norm_text(v)]
        title = text_cells[0] if text_cells else ""
        numbers = [_parse_number(v) for v in row]
        nums = [n for n in numbers if n is not None]

        title_lower = title.lower()
        is_total_row = title_lower.startswith("итого") or title_lower.startswith("всего") or title_lower.startswith("total")
        if title and _is_aggregate_or_tax_row_label(title):
            continue
        is_header_candidate = bool(title) and not nums
        if is_header_candidate:
            if any(token in title_lower for token in _ROW_HEADER_HINTS):
                continue
            ensure_block(title)
            continue

        block = current_block or ensure_block(fallback_block_name)
        qty: float | None = None
        unit: float | None = None
        amount: float | None = None

        if len(nums) >= 3:
            qty, unit, amount = nums[-3], nums[-2], nums[-1]
        elif len(nums) == 2:
            first, second = nums
            if abs(first - round(first)) < 1e-6 and 0 < first <= 200:
                qty = first
                amount = second
            else:
                unit = first
                amount = second
        elif len(nums) == 1:
            amount = nums[0]

        if amount is None and qty is not None and unit is not None:
            amount = unit if qty == 0 else qty * unit

        if is_total_row:
            if amount is not None:
                block["explicit_total"] = float(amount)
            continue

        item_title = title or f"Позиция {len(block['items']) + 1}"
        if amount is None and qty is None and unit is None:
            continue

        block["items"].append(
            {
                "title": item_title,
                "qty": None if qty is None else float(qty),
                "unit": None if unit is None else float(unit),
                "amount": None if amount is None else float(amount),
            }
        )

    finalized: list[dict] = []
    for block in blocks:
        if not block["items"]:
            continue
        computed_total = sum(float(it["amount"] or 0.0) for it in block["items"])
        finalized.append(
            {
                "title": block["title"],
                "items": block["items"],
                "total": max(0.0, float(computed_total)),
            }
        )
    return finalized


def _detect_estimate_profile(rows: list[list[object]]) -> str:
    joined = " ".join(
        _norm_text(cell).lower()
        for row in rows[:180]
        for cell in row
        if _norm_text(cell)
    )
    if "общая застройка по зонам" in joined:
        return "zones_v1"
    if "ценовое предложение выставочного стенда" in joined:
        return "sections_v1"
    if "итого по разделу" in joined:
        return "sections_v1"
    return "generic"


def _parse_rows_zones_v1(rows: list[list[object]], fallback_block_name: str) -> tuple[list[dict], list[str]]:
    blocks: list[dict] = []
    warnings: list[str] = []
    current_block: dict | None = None

    def ensure_block(name: str) -> dict:
        nonlocal current_block
        title = _norm_text(name) or fallback_block_name
        if current_block and current_block["title"] == title:
            return current_block
        block = {"title": title, "items": [], "explicit_total": None}
        blocks.append(block)
        current_block = block
        return block

    for row_idx, row in enumerate(rows, start=1):
        text_cells = [_norm_text(v) for v in row if isinstance(v, str) and _norm_text(v)]
        if not text_cells and not any(_parse_number(v) is not None for v in row):
            continue
        first_text = text_cells[0] if text_cells else ""
        first_lower = first_text.lower()
        if first_lower in _ROW_HEADER_HINTS or first_lower.startswith("наименование"):
            continue
        if first_text and _is_aggregate_or_tax_row_label(first_text):
            continue

        nums_raw = [_parse_number(v) for v in row]
        nums = [n for n in nums_raw if n is not None]
        is_total = first_lower.startswith("итого") or first_lower.startswith("всего")

        if len(text_cells) == 1 and not nums and not is_total:
            ensure_block(first_text)
            continue

        block = current_block or ensure_block(fallback_block_name)
        if is_total:
            if nums:
                block["explicit_total"] = float(nums[-1])
            continue

        title = first_text or f"Позиция {len(block['items']) + 1}"
        qty = unit = amount = None
        if len(nums) >= 3:
            qty, unit, amount = nums[-3], nums[-2], nums[-1]
        elif len(nums) == 2:
            qty, amount = nums[0], nums[1]
        elif len(nums) == 1:
            amount = nums[0]
        if amount is None and qty is not None and unit is not None:
            amount = unit if qty == 0 else qty * unit
        if amount is None and qty is None and unit is None:
            continue
        block["items"].append({"title": title, "qty": qty, "unit": unit, "amount": amount})

        if amount is None:
            warnings.append(f"ROW_{row_idx}: SUM_MISSING")

    finalized: list[dict] = []
    for block in blocks:
        if not block["items"]:
            continue
        computed_total = sum(float(it["amount"] or 0.0) for it in block["items"])
        explicit_total = block.get("explicit_total")
        if explicit_total is not None and abs(float(explicit_total) - float(computed_total)) > 0.01:
            warnings.append(f"BLOCK_TOTAL_MISMATCH:{block['title']}")
        finalized.append({"title": block["title"], "items": block["items"], "total": max(0.0, computed_total)})
    return finalized, warnings


def _is_section_row(first: str, second: str, nums: list[float]) -> bool:
    if not second:
        return False
    if re.fullmatch(r"\d+(\.)?", first.strip()):
        return True
    if re.fullmatch(r"\d+(\.\d+)?", first.strip()) and len(nums) <= 1:
        return True
    return False


def _parse_rows_sections_v1(rows: list[list[object]], fallback_block_name: str) -> tuple[list[dict], list[str]]:
    blocks: list[dict] = []
    warnings: list[str] = []
    current_block: dict | None = None
    table_started = False

    def ensure_block(name: str) -> dict:
        nonlocal current_block
        title = _norm_text(name) or fallback_block_name
        if current_block and current_block["title"] == title:
            return current_block
        block = {"title": title, "items": [], "explicit_total": None}
        blocks.append(block)
        current_block = block
        return block

    for row_idx, row in enumerate(rows, start=1):
        cells = [_norm_text(v) for v in row]
        text_cells = [c for c in cells if c]
        if not text_cells and not any(_parse_number(v) is not None for v in row):
            continue

        first = text_cells[0] if text_cells else ""
        second = text_cells[1] if len(text_cells) > 1 else ""
        first_lower = first.lower()
        row_joined = " ".join(text_cells).lower()
        if not table_started:
            if (
                "наименование" in row_joined
                and ("стоимость" in row_joined or "общая стоимость" in row_joined)
            ):
                table_started = True
            continue
        if first_lower in _ROW_HEADER_HINTS or first_lower.startswith("наименование"):
            continue
        if first and _is_aggregate_or_tax_row_label(first):
            continue

        nums_raw = [_parse_number(v) for v in row]
        nums = [n for n in nums_raw if n is not None]
        if _is_section_row(first, second, nums):
            ensure_block(second)
            continue

        is_total = "итого по разделу" in first_lower or first_lower.startswith("итого") or first_lower.startswith("всего")
        block = current_block or ensure_block(fallback_block_name)
        if is_total:
            if nums:
                block["explicit_total"] = float(nums[-1])
            continue

        title = first
        if re.fullmatch(r"\d+(\.)?", title) and second:
            title = second
        if not title:
            title = f"Позиция {len(block['items']) + 1}"

        qty = unit = amount = None
        if len(nums) >= 3:
            qty, unit, amount = nums[-3], nums[-2], nums[-1]
        elif len(nums) == 2:
            qty, amount = nums[0], nums[1]
        elif len(nums) == 1:
            amount = nums[0]
        if amount is None and qty is not None and unit is not None:
            amount = unit if qty == 0 else qty * unit
        if amount is None and qty is None and unit is None:
            continue
        block["items"].append({"title": title, "qty": qty, "unit": unit, "amount": amount})

        if amount is None:
            warnings.append(f"ROW_{row_idx}: SUM_MISSING")

    finalized: list[dict] = []
    for block in blocks:
        if not block["items"]:
            continue
        computed_total = sum(float(it["amount"] or 0.0) for it in block["items"])
        explicit_total = block.get("explicit_total")
        if explicit_total is not None and abs(float(explicit_total) - float(computed_total)) > 0.01:
            warnings.append(f"BLOCK_TOTAL_MISMATCH:{block['title']}")
        finalized.append({"title": block["title"], "items": block["items"], "total": max(0.0, computed_total)})
    return finalized, warnings


def _parse_contractor_estimate(rows: list[list[object]], fallback_name: str) -> tuple[str, list[dict], list[str]]:
    profile = _detect_estimate_profile(rows)
    warnings: list[str] = []
    if profile == "zones_v1":
        parsed_blocks, warnings = _parse_rows_zones_v1(rows, fallback_name)
    elif profile == "sections_v1":
        parsed_blocks, warnings = _parse_rows_sections_v1(rows, fallback_name)
    else:
        parsed_blocks = _parse_imported_estimate_rows(rows, fallback_name)
    return profile, parsed_blocks, warnings


def _apply_block_overrides(parsed_blocks: list[dict], overrides_raw: Optional[str]) -> tuple[list[dict], list[str]]:
    if not overrides_raw:
        return parsed_blocks, []

    warnings: list[str] = []
    try:
        payload = json.loads(overrides_raw)
    except Exception:
        return parsed_blocks, ["OVERRIDES_JSON_INVALID"]
    if not isinstance(payload, list):
        return parsed_blocks, ["OVERRIDES_FORMAT_INVALID"]

    include_map: dict[int, bool] = {}
    rename_map: dict[int, str] = {}
    for entry in payload:
        if not isinstance(entry, dict):
            continue
        raw_index = entry.get("block_index")
        if not isinstance(raw_index, int):
            continue
        if raw_index < 0 or raw_index >= len(parsed_blocks):
            warnings.append(f"OVERRIDE_INDEX_OUT_OF_RANGE:{raw_index}")
            continue
        if "include" in entry:
            include_map[raw_index] = bool(entry.get("include"))
        raw_title = _norm_text(entry.get("title"))
        if raw_title:
            rename_map[raw_index] = raw_title

    filtered: list[dict] = []
    for idx, block in enumerate(parsed_blocks):
        if idx in include_map and not include_map[idx]:
            continue
        next_block = dict(block)
        if idx in rename_map:
            next_block["title"] = rename_map[idx]
        filtered.append(next_block)
    return filtered, warnings


def _norm_key(value: object) -> str:
    text = _norm_text(value).lower().replace("ё", "е")
    text = re.sub(r"^\d+([.)]\d+)*\s*", "", text)
    text = re.sub(r"[^a-z0-9а-я]+", "", text)
    return text


def _row_amount_from_parsed(row: dict) -> float:
    qty = row.get("qty")
    unit = row.get("unit")
    amount = row.get("amount")
    if qty is not None and unit is not None and (amount is None):
        return float(unit) if float(qty) == 0 else float(qty) * float(unit)
    if amount is not None:
        return max(0.0, float(amount))
    if qty is not None and unit is None:
        return max(0.0, float(qty))
    return 0.0


def _item_current_amount(item: ExpenseItem) -> float:
    if item.mode == ItemMode.QTY_PRICE:
        qty = float(item.qty or 0.0)
        unit = float(item.unit_price_base or 0.0)
        return unit if qty == 0 else qty * unit
    return float(item.base_total or 0.0)


def _load_group_top_level_with_children(
    db: Session,
    project_id: int,
    group_id: int,
) -> tuple[dict[str, tuple[ExpenseItem, list[ExpenseItem]]], list[ExpenseItem]]:
    items = db.execute(
        select(ExpenseItem).where(
            ExpenseItem.project_id == project_id,
            ExpenseItem.group_id == group_id,
        ).order_by(ExpenseItem.id.asc())
    ).scalars().all()
    by_parent: dict[int, list[ExpenseItem]] = {}
    all_ids = {int(it.id) for it in items}
    for it in items:
        if it.parent_item_id is None:
            continue
        parent_id = int(it.parent_item_id)
        by_parent.setdefault(parent_id, []).append(it)
    top_level = [it for it in items if it.parent_item_id is None or int(it.parent_item_id) not in all_ids]
    blocks_by_key: dict[str, tuple[ExpenseItem, list[ExpenseItem]]] = {}
    for parent in top_level:
        children = by_parent.get(int(parent.id), [])
        if not children:
            continue
        blocks_by_key[_norm_key(parent.title)] = (parent, children)
    return blocks_by_key, top_level


def _build_preview_diff_rows(
    parsed_blocks: list[dict],
    existing_blocks_by_key: dict[str, tuple[ExpenseItem, list[ExpenseItem]]],
) -> list[dict]:
    diff_rows: list[dict] = []
    for block in parsed_blocks:
        block_title = str(block.get("title") or "Блок")
        block_key = _norm_key(block_title)
        existing = existing_blocks_by_key.get(block_key)
        existing_children_by_key: dict[str, ExpenseItem] = {}
        if existing:
            _, existing_children = existing
            for child in existing_children:
                existing_children_by_key[_norm_key(child.title)] = child
        matched_existing_keys: set[str] = set()
        for row in (block.get("items") or []):
            row_title = str(row.get("title") or "Позиция")
            row_key = _norm_key(row_title)
            new_amount = _row_amount_from_parsed(row)
            existing_child = existing_children_by_key.get(row_key)
            if existing_child is None:
                diff_rows.append(
                    {
                        "block_title": block_title,
                        "row_title": row_title,
                        "status": "new",
                        "old_amount": None,
                        "new_amount": new_amount,
                    }
                )
                continue
            matched_existing_keys.add(row_key)
            old_amount = _item_current_amount(existing_child)
            if abs(new_amount - old_amount) > 0.01:
                diff_rows.append(
                    {
                        "block_title": block_title,
                        "row_title": row_title,
                        "status": "changed",
                        "old_amount": old_amount,
                        "new_amount": new_amount,
                    }
                )
            else:
                diff_rows.append(
                    {
                        "block_title": block_title,
                        "row_title": row_title,
                        "status": "unchanged",
                        "old_amount": old_amount,
                        "new_amount": new_amount,
                    }
                )
        for key, existing_child in existing_children_by_key.items():
            if key in matched_existing_keys:
                continue
            diff_rows.append(
                {
                    "block_title": block_title,
                    "row_title": str(existing_child.title or "Позиция"),
                    "status": "removed",
                    "old_amount": _item_current_amount(existing_child),
                    "new_amount": None,
                }
            )
    return diff_rows

def _get_project_or_404(db: Session, project_id: int) -> Project:
    p = db.get(Project, project_id)
    if not p:
        raise HTTPException(404, "PROJECT_NOT_FOUND")
    return p

def _get_group_or_404(db: Session, project_id: int, group_id: int) -> ExpenseGroup:
    g = db.get(ExpenseGroup, group_id)
    if not g or g.project_id != project_id:
        raise HTTPException(404, "GROUP_NOT_FOUND")
    return g

def _get_item_or_404(db: Session, project_id: int, item_id: int) -> ExpenseItem:
    it = db.get(ExpenseItem, item_id)
    if not it or it.project_id != project_id:
        raise HTTPException(404, "ITEM_NOT_FOUND")
    return it


def _delete_group_with_dependencies(db: Session, project_id: int, group_id: int) -> None:
    g = _get_group_or_404(db, project_id, group_id)
    try:
        item_ids = db.execute(
            select(ExpenseItem.id).where(
                ExpenseItem.project_id == project_id,
                ExpenseItem.group_id == group_id,
            )
        ).scalars().all()
        if item_ids:
            db.execute(
                delete(ClientBillingAdjustment).where(
                    ClientBillingAdjustment.expense_item_id.in_(item_ids)
                )
            )
            db.execute(
                delete(ExpenseItem).where(
                    ExpenseItem.project_id == project_id,
                    ExpenseItem.group_id == group_id,
                )
            )
        db.delete(g)
        db.commit()
    except SQLAlchemyError as exc:
        db.rollback()
        raise HTTPException(500, f"GROUP_DELETE_FAILED: {exc.__class__.__name__}")

def _parse_mode(mode_raw: str) -> ItemMode:
    try:
        return ItemMode(mode_raw)
    except ValueError as exc:
        raise HTTPException(422, "ITEM_MODE_INVALID") from exc

def _parse_adjustment_type(value: str) -> AdjustmentType:
    try:
        return AdjustmentType(value)
    except ValueError as exc:
        raise HTTPException(422, "ADJUSTMENT_TYPE_INVALID") from exc

def _refresh_item_calculated_base(item: ExpenseItem) -> None:
    if item.mode == ItemMode.QTY_PRICE:
        if item.qty is None or item.unit_price_base is None:
            raise HTTPException(422, "QTY_PRICE_REQUIRES_QTY_AND_UNIT_PRICE")
        qty = float(item.qty)
        unit = float(item.unit_price_base)
        item.base_total = unit if qty == 0 else qty * unit


def _validate_parent_item(
    db: Session,
    project_id: int,
    group_id: int,
    parent_item_id: int | None,
    current_item_id: int | None = None,
) -> None:
    if parent_item_id is None:
        return
    if current_item_id is not None and parent_item_id == current_item_id:
        raise HTTPException(422, "PARENT_ITEM_SELF_REF")
    parent = db.get(ExpenseItem, parent_item_id)
    if not parent or parent.project_id != project_id:
        raise HTTPException(422, "PARENT_ITEM_NOT_FOUND")
    if parent.group_id != group_id:
        raise HTTPException(422, "PARENT_ITEM_GROUP_MISMATCH")
    # One nesting level: only top-level rows can be parents.
    if parent.parent_item_id is not None:
        raise HTTPException(422, "PARENT_ITEM_MUST_BE_TOP_LEVEL")
    if current_item_id is not None:
        has_children = db.execute(
            select(ExpenseItem.id).where(ExpenseItem.parent_item_id == current_item_id).limit(1)
        ).first() is not None
        if has_children:
            raise HTTPException(422, "ITEM_WITH_SUBITEMS_CANNOT_BE_SUBITEM")


def _load_discount_adjustment_for_item(db: Session, item_id: int) -> ClientBillingAdjustment | None:
    return db.execute(
        select(ClientBillingAdjustment).where(ClientBillingAdjustment.expense_item_id == item_id)
    ).scalar_one_or_none()


def _apply_item_discount(
    db: Session,
    item_id: int,
    discount_enabled: bool,
    discount_amount: float,
) -> None:
    adj = _load_discount_adjustment_for_item(db, item_id)
    if not adj:
        adj = ClientBillingAdjustment(
            expense_item_id=item_id,
            unit_price_full=0.0,
            unit_price_billable=0.0,
            adjustment_type=AdjustmentType.DISCOUNT,
            reason="",
        )
        db.add(adj)
    adj.discount_enabled = bool(discount_enabled)
    adj.discount_amount = float(discount_amount or 0.0)


def _attach_item_discounts(db: Session, project_id: int, items: list[ExpenseItem]) -> list[ExpenseItem]:
    if not items:
        return items
    adjustments = db.execute(
        select(ClientBillingAdjustment).join(ExpenseItem, ExpenseItem.id == ClientBillingAdjustment.expense_item_id).where(
            ExpenseItem.project_id == project_id
        )
    ).scalars().all()
    by_item_id = {int(adj.expense_item_id): adj for adj in adjustments}
    for it in items:
        adj = by_item_id.get(int(it.id))
        setattr(it, "discount_enabled", bool(adj.discount_enabled) if adj else False)
        setattr(it, "discount_amount", float(adj.discount_amount or 0.0) if adj else 0.0)
    return items

@router.get("", response_model=list[ProjectOut])
def list_projects(db: Session = Depends(get_db)):
    return db.execute(
        select(Project).order_by(Project.is_paused.asc(), Project.sort_order.desc(), Project.id.desc())
    ).scalars().all()


@router.post("/reorder")
def reorder_projects(payload: ProjectReorderIn, db: Session = Depends(get_db)):
    ids = [int(v) for v in payload.project_ids]
    if not ids:
        return {"updated": 0}
    if len(set(ids)) != len(ids):
        raise HTTPException(422, "PROJECT_REORDER_DUPLICATES")

    all_ids = db.execute(
        select(Project.id).order_by(Project.sort_order.desc(), Project.id.desc())
    ).scalars().all()
    all_ids_set = set(int(v) for v in all_ids)
    if any(pid not in all_ids_set for pid in ids):
        raise HTTPException(422, "PROJECT_REORDER_UNKNOWN_ID")

    pinned = set(ids)
    remainder = [int(pid) for pid in all_ids if int(pid) not in pinned]
    ordered_ids = ids + remainder
    total = len(ordered_ids)
    for idx, project_id in enumerate(ordered_ids):
        db.execute(
            update(Project)
            .where(Project.id == project_id)
            .values(sort_order=total - idx)
        )
    db.commit()
    return {"updated": total}

@router.post("", response_model=ProjectOut)
def create_project(payload: ProjectCreate, db: Session = Depends(get_db)):
    max_sort_order = db.execute(select(func.coalesce(func.max(Project.sort_order), 0))).scalar_one()
    p = Project(
        title=payload.title,
        client_name=payload.client_name,
        client_email=payload.client_email,
        client_phone=payload.client_phone,
        google_drive_url=payload.google_drive_url,
        google_drive_folder=payload.google_drive_folder,
        card_image_data=payload.card_image_data,
        agency_fee_percent=payload.agency_fee_percent,
        agency_fee_include_in_estimate=payload.agency_fee_include_in_estimate,
        sort_order=int(max_sort_order) + 1,
        is_paused=False,
        project_price_total=payload.project_price_total,
        expected_from_client_total=payload.expected_from_client_total,
        closed_at=payload.closed_at,
    )
    db.add(p)
    db.commit()
    db.refresh(p)
    return p

@router.get("/{project_id}", response_model=ProjectOut)
def get_project(project_id: int, db: Session = Depends(get_db)):
    return _get_project_or_404(db, project_id)

@router.patch("/{project_id}", response_model=ProjectOut)
def update_project(project_id: int, payload: ProjectUpdate, db: Session = Depends(get_db)):
    p = _get_project_or_404(db, project_id)

    data = payload.model_dump(exclude_unset=True)
    pause_update = "is_paused" in data and data["is_paused"] is not None
    next_paused = bool(data.pop("is_paused")) if pause_update else bool(p.is_paused)

    for k, v in data.items():
        setattr(p, k, v)

    if pause_update and next_paused != bool(p.is_paused):
        p.is_paused = next_paused
        if not next_paused:
            max_sort_order = db.execute(
                select(func.coalesce(func.max(Project.sort_order), 0)).where(
                    Project.is_paused.is_(False),
                    Project.id != p.id,
                )
            ).scalar_one()
            p.sort_order = int(max_sort_order) + 1

    db.commit()
    db.refresh(p)
    return p

@router.delete("/{project_id}")
def delete_project(project_id: int, db: Session = Depends(get_db)):
    p = _get_project_or_404(db, project_id)
    db.delete(p)
    db.commit()
    return {"deleted": True}

@router.get("/{project_id}/computed", response_model=ProjectComputed)
def project_computed(project_id: int, db: Session = Depends(get_db)):
    _get_project_or_404(db, project_id)
    return compute_project_financials(db, project_id)

@router.get("/{project_id}/groups", response_model=list[GroupOut])
def list_groups(project_id: int, db: Session = Depends(get_db)):
    _get_project_or_404(db, project_id)
    return db.execute(select(ExpenseGroup).where(ExpenseGroup.project_id == project_id).order_by(ExpenseGroup.sort_order.asc())).scalars().all()

@router.post("/{project_id}/groups", response_model=GroupOut)
def create_group(project_id: int, payload: GroupCreate, db: Session = Depends(get_db)):
    _get_project_or_404(db, project_id)
    g = ExpenseGroup(project_id=project_id, name=payload.name, sort_order=payload.sort_order)
    db.add(g)
    db.commit()
    db.refresh(g)
    return g

@router.patch("/{project_id}/groups/{group_id}", response_model=GroupOut)
def update_group(project_id: int, group_id: int, payload: GroupUpdate, db: Session = Depends(get_db)):
    g = _get_group_or_404(db, project_id, group_id)
    data = payload.model_dump(exclude_unset=True)

    if "name" in data and data["name"] is not None:
        name = str(data["name"]).strip()
        if not name:
            raise HTTPException(422, "GROUP_NAME_EMPTY")
        g.name = name

    if "sort_order" in data and data["sort_order"] is not None:
        g.sort_order = int(data["sort_order"])

    db.commit()
    db.refresh(g)
    return g

@router.delete("/{project_id}/groups/{group_id}")
def delete_group(project_id: int, group_id: int, db: Session = Depends(get_db)):
    _delete_group_with_dependencies(db, project_id, group_id)
    return {"deleted": True}


@router.post("/{project_id}/groups/{group_id}/delete")
def delete_group_post(project_id: int, group_id: int, db: Session = Depends(get_db)):
    _delete_group_with_dependencies(db, project_id, group_id)
    return {"deleted": True}


@router.post(
    "/{project_id}/groups/{group_id}/contractor-estimate/preview",
    response_model=ContractorEstimatePreviewOut,
)
async def preview_contractor_estimate(
    project_id: int,
    group_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    _get_project_or_404(db, project_id)
    group = _get_group_or_404(db, project_id, group_id)

    filename = _norm_text(file.filename or "")
    if not filename:
        raise HTTPException(422, "ESTIMATE_FILE_NAME_REQUIRED")
    content = await file.read()
    if not content:
        raise HTTPException(422, "ESTIMATE_FILE_EMPTY")

    table_rows = _load_table_rows_from_file(filename, content)
    profile, parsed_blocks, warnings = _parse_contractor_estimate(
        table_rows,
        fallback_name=f"Импорт ({group.name})",
    )
    if not parsed_blocks:
        raise HTTPException(422, "ESTIMATE_PARSE_EMPTY")
    existing_blocks_by_key, _ = _load_group_top_level_with_children(db, project_id, group_id)
    diff_rows = _build_preview_diff_rows(parsed_blocks, existing_blocks_by_key)

    preview_blocks = []
    total_items = 0
    for idx, block in enumerate(parsed_blocks):
        block_items = block.get("items") or []
        total_items += len(block_items)
        preview_blocks.append(
            {
                "block_index": idx,
                "title": str(block.get("title") or "Блок"),
                "items": len(block_items),
                "total": float(block.get("total") or 0.0),
                "sample_rows": [str(it.get("title") or "") for it in block_items[:5]],
            }
        )
    total_items = sum(len((b.get("items") or [])) for b in parsed_blocks)
    return {
        "ok": True,
        "profile": profile,
        "blocks": len(parsed_blocks),
        "items": total_items,
        "warnings": warnings[:100],
        "preview_blocks": preview_blocks,
        "diff_rows": diff_rows[:400],
    }


@router.post(
    "/{project_id}/groups/{group_id}/contractor-estimate/import",
    response_model=ContractorEstimateImportOut,
)
async def import_contractor_estimate(
    project_id: int,
    group_id: int,
    file: UploadFile = File(...),
    overrides: Optional[str] = Form(default=None),
    planned_pay_date: str = Form(...),
    db: Session = Depends(get_db),
):
    _get_project_or_404(db, project_id)
    group = _get_group_or_404(db, project_id, group_id)

    filename = _norm_text(file.filename or "")
    if not filename:
        raise HTTPException(422, "ESTIMATE_FILE_NAME_REQUIRED")
    content = await file.read()
    if not content:
        raise HTTPException(422, "ESTIMATE_FILE_EMPTY")
    pay_date_raw = _norm_text(planned_pay_date)
    if not pay_date_raw:
        raise HTTPException(422, "PLANNED_PAY_DATE_REQUIRED")
    try:
        pay_date = date.fromisoformat(pay_date_raw)
    except ValueError:
        raise HTTPException(422, "PLANNED_PAY_DATE_INVALID")

    table_rows = _load_table_rows_from_file(filename, content)
    profile, parsed_blocks, warnings = _parse_contractor_estimate(
        table_rows,
        fallback_name=f"Импорт ({group.name})",
    )
    parsed_blocks, override_warnings = _apply_block_overrides(parsed_blocks, overrides)
    if override_warnings:
        warnings = [*warnings, *override_warnings]
    if not parsed_blocks:
        raise HTTPException(422, "ESTIMATE_PARSE_EMPTY")

    existing_blocks_by_key, _ = _load_group_top_level_with_children(db, project_id, group_id)
    imported_blocks = 0
    imported_items = 0
    created_parent_item_ids: list[int] = []
    for block in parsed_blocks:
        block_title = str(block.get("title") or "Блок")
        block_key = _norm_key(block_title)
        existing = existing_blocks_by_key.get(block_key)
        if existing is None:
            parent = ExpenseItem(
                stable_item_id=gen_stable_id("item"),
                project_id=project_id,
                group_id=group_id,
                parent_item_id=None,
                title=block_title,
                mode=ItemMode.SINGLE_TOTAL,
                qty=None,
                unit_price_base=None,
                base_total=float(block["total"]),
                include_in_estimate=True,
                extra_profit_enabled=False,
                extra_profit_amount=0.0,
                planned_pay_date=pay_date,
            )
            db.add(parent)
            db.flush()
            created_parent_item_ids.append(int(parent.id))
            existing_children: list[ExpenseItem] = []
        else:
            parent, existing_children = existing
            parent.title = block_title
            parent.mode = ItemMode.SINGLE_TOTAL
            parent.qty = None
            parent.unit_price_base = None
            parent.base_total = max(0.0, float(block.get("total") or 0.0))
            parent.include_in_estimate = True
            parent.planned_pay_date = pay_date
        imported_blocks += 1

        existing_children_by_key = {_norm_key(ch.title): ch for ch in existing_children}
        matched_existing_child_ids: set[int] = set()
        for row in (block.get("items") or []):
            qty = row.get("qty")
            unit = row.get("unit")
            amount = _row_amount_from_parsed(row)
            mode = ItemMode.SINGLE_TOTAL
            base_total = float(amount or 0.0)
            out_qty: float | None = None
            out_unit: float | None = None

            if qty is not None and unit is not None:
                computed = float(unit) if float(qty) == 0 else float(qty) * float(unit)
                if abs(float(amount) - computed) <= 0.01:
                    mode = ItemMode.QTY_PRICE
                    out_qty = float(qty)
                    out_unit = float(unit)
                    base_total = computed
                else:
                    mode = ItemMode.SINGLE_TOTAL
                    base_total = float(amount)
            row_title = str(row.get("title") or "Позиция")
            row_key = _norm_key(row_title)
            existing_child = existing_children_by_key.get(row_key)

            if existing_child is None:
                child = ExpenseItem(
                    stable_item_id=gen_stable_id("item"),
                    project_id=project_id,
                    group_id=group_id,
                    parent_item_id=int(parent.id),
                    title=row_title,
                    mode=mode,
                    qty=out_qty,
                    unit_price_base=out_unit,
                    base_total=max(0.0, float(base_total)),
                    include_in_estimate=False,
                    extra_profit_enabled=False,
                    extra_profit_amount=0.0,
                    planned_pay_date=pay_date,
                )
                _refresh_item_calculated_base(child)
                db.add(child)
            else:
                old_amount = _item_current_amount(existing_child)
                existing_child.title = row_title
                existing_child.mode = mode
                existing_child.qty = out_qty
                existing_child.unit_price_base = out_unit
                existing_child.base_total = max(0.0, float(base_total))
                existing_child.include_in_estimate = False
                existing_child.planned_pay_date = pay_date
                if mode == ItemMode.QTY_PRICE:
                    _refresh_item_calculated_base(existing_child)
                if base_total > old_amount + 0.01:
                    existing_child.extra_profit_enabled = True
                    existing_child.extra_profit_amount = float(existing_child.extra_profit_amount or 0.0) + (base_total - old_amount)
                matched_existing_child_ids.add(int(existing_child.id))
            imported_items += 1
        for existing_child in existing_children:
            if int(existing_child.id) in matched_existing_child_ids:
                continue
            existing_child.base_total = 0.0
            existing_child.mode = ItemMode.SINGLE_TOTAL
            existing_child.qty = None
            existing_child.unit_price_base = None
            existing_child.include_in_estimate = False
            existing_child.extra_profit_enabled = False
            existing_child.extra_profit_amount = 0.0
            existing_child.planned_pay_date = pay_date

    db.commit()
    return {
        "ok": True,
        "imported_blocks": imported_blocks,
        "imported_items": imported_items,
        "created_parent_item_ids": created_parent_item_ids,
        "profile": profile,
        "warnings": warnings[:100],
    }

@router.get("/{project_id}/items", response_model=list[ItemOut])
def list_items(project_id: int, db: Session = Depends(get_db)):
    _get_project_or_404(db, project_id)
    items = db.execute(
        select(ExpenseItem).where(ExpenseItem.project_id == project_id).order_by(ExpenseItem.group_id.asc(), ExpenseItem.id.asc())
    ).scalars().all()
    return _attach_item_discounts(db, project_id, items)

@router.post("/{project_id}/items", response_model=ItemOut)
def create_item(project_id: int, payload: ItemCreate, db: Session = Depends(get_db)):
    _get_project_or_404(db, project_id)
    _get_group_or_404(db, project_id, payload.group_id)
    mode = _parse_mode(payload.mode)
    _validate_parent_item(
        db=db,
        project_id=project_id,
        group_id=payload.group_id,
        parent_item_id=payload.parent_item_id,
    )

    it = ExpenseItem(
        stable_item_id=gen_stable_id("item"),
        project_id=project_id,
        group_id=payload.group_id,
        parent_item_id=payload.parent_item_id,
        title=payload.title,
        mode=mode,
        qty=payload.qty,
        unit_price_base=payload.unit_price_base,
        base_total=payload.base_total,
        include_in_estimate=payload.include_in_estimate,
        extra_profit_enabled=payload.extra_profit_enabled,
        extra_profit_amount=payload.extra_profit_amount,
        planned_pay_date=payload.planned_pay_date,
    )
    _refresh_item_calculated_base(it)
    db.add(it)
    db.flush()
    _apply_item_discount(
        db=db,
        item_id=int(it.id),
        discount_enabled=bool(payload.discount_enabled),
        discount_amount=float(payload.discount_amount or 0.0),
    )
    db.commit()
    db.refresh(it)
    _attach_item_discounts(db, project_id, [it])
    return it

@router.patch("/{project_id}/items/{item_id}", response_model=ItemOut)
def update_item(project_id: int, item_id: int, payload: ItemUpdate, db: Session = Depends(get_db)):
    it = _get_item_or_404(db, project_id, item_id)
    data = payload.model_dump(exclude_unset=True)
    discount_enabled = data.pop("discount_enabled", None)
    discount_amount = data.pop("discount_amount", None)
    has_children = db.execute(
        select(ExpenseItem.id).where(ExpenseItem.parent_item_id == item_id).limit(1)
    ).first() is not None

    if "group_id" in data and data["group_id"] is not None:
        _get_group_or_404(db, project_id, data["group_id"])
        if has_children and int(data["group_id"]) != it.group_id:
            raise HTTPException(422, "ITEM_WITH_SUBITEMS_CANNOT_CHANGE_GROUP")

    if "mode" in data and data["mode"] is not None:
        data["mode"] = _parse_mode(data["mode"])

    target_group_id = int(data["group_id"]) if data.get("group_id") is not None else int(it.group_id)
    target_parent_item_id = data["parent_item_id"] if "parent_item_id" in data else it.parent_item_id
    _validate_parent_item(
        db=db,
        project_id=project_id,
        group_id=target_group_id,
        parent_item_id=target_parent_item_id,
        current_item_id=item_id,
    )

    for k, v in data.items():
        setattr(it, k, v)

    _refresh_item_calculated_base(it)
    if discount_enabled is not None or discount_amount is not None:
        current_adj = _load_discount_adjustment_for_item(db, item_id)
        resolved_enabled = bool(discount_enabled) if discount_enabled is not None else bool(current_adj.discount_enabled) if current_adj else False
        resolved_amount = float(discount_amount) if discount_amount is not None else float(current_adj.discount_amount or 0.0) if current_adj else 0.0
        _apply_item_discount(
            db=db,
            item_id=item_id,
            discount_enabled=resolved_enabled,
            discount_amount=resolved_amount,
        )
    db.commit()
    db.refresh(it)
    _attach_item_discounts(db, project_id, [it])
    return it

@router.delete("/{project_id}/items/{item_id}")
def delete_item(project_id: int, item_id: int, db: Session = Depends(get_db)):
    it = _get_item_or_404(db, project_id, item_id)
    subitems = db.execute(
        select(ExpenseItem).where(ExpenseItem.parent_item_id == item_id)
    ).scalars().all()
    for subitem in subitems:
        db.delete(subitem)
    db.delete(it)
    db.commit()
    return {"deleted": True}

@router.get("/{project_id}/items/{item_id}/adjustment", response_model=BillingAdjustmentOut)
def get_item_adjustment(project_id: int, item_id: int, db: Session = Depends(get_db)):
    _get_item_or_404(db, project_id, item_id)
    adj = db.execute(
        select(ClientBillingAdjustment).where(ClientBillingAdjustment.expense_item_id == item_id)
    ).scalar_one_or_none()
    if not adj:
        raise HTTPException(404, "ADJUSTMENT_NOT_FOUND")
    return adj

@router.put("/{project_id}/items/{item_id}/adjustment", response_model=BillingAdjustmentOut)
def upsert_item_adjustment(project_id: int, item_id: int, payload: BillingAdjustmentUpsert, db: Session = Depends(get_db)):
    _get_item_or_404(db, project_id, item_id)
    adj = db.execute(
        select(ClientBillingAdjustment).where(ClientBillingAdjustment.expense_item_id == item_id)
    ).scalar_one_or_none()
    adj_type = _parse_adjustment_type(payload.adjustment_type)

    if not adj:
        adj = ClientBillingAdjustment(
            expense_item_id=item_id,
            unit_price_full=payload.unit_price_full,
            unit_price_billable=payload.unit_price_billable,
            adjustment_type=adj_type,
            reason=payload.reason,
        )
        db.add(adj)
    else:
        adj.unit_price_full = payload.unit_price_full
        adj.unit_price_billable = payload.unit_price_billable
        adj.adjustment_type = adj_type
        adj.reason = payload.reason

    db.commit()
    db.refresh(adj)
    return adj

@router.delete("/{project_id}/items/{item_id}/adjustment")
def delete_item_adjustment(project_id: int, item_id: int, db: Session = Depends(get_db)):
    _get_item_or_404(db, project_id, item_id)
    adj = db.execute(
        select(ClientBillingAdjustment).where(ClientBillingAdjustment.expense_item_id == item_id)
    ).scalar_one_or_none()
    if not adj:
        raise HTTPException(404, "ADJUSTMENT_NOT_FOUND")
    db.delete(adj)
    db.commit()
    return {"deleted": True}

@router.get("/{project_id}/payments/plan", response_model=list[PaymentPlanOut])
def list_payments_plan(project_id: int, db: Session = Depends(get_db)):
    _get_project_or_404(db, project_id)
    today = date.today()
    return db.execute(
        select(ClientPaymentsPlan).where(
            ClientPaymentsPlan.project_id == project_id,
            ClientPaymentsPlan.pay_date > today,
        ).order_by(ClientPaymentsPlan.pay_date.asc(), ClientPaymentsPlan.id.asc())
    ).scalars().all()

@router.post("/{project_id}/payments/plan", response_model=PaymentPlanOut)
def create_payment_plan(project_id: int, payload: PaymentPlanCreate, db: Session = Depends(get_db)):
    _get_project_or_404(db, project_id)
    rec = ClientPaymentsPlan(
        stable_pay_id=gen_stable_id("pay"),
        project_id=project_id,
        pay_date=payload.pay_date,
        amount=payload.amount,
        note=payload.note,
    )
    db.add(rec)
    db.commit()
    db.refresh(rec)
    return rec

@router.patch("/{project_id}/payments/plan/{pay_id}", response_model=PaymentPlanOut)
def update_payment_plan(project_id: int, pay_id: int, payload: PaymentPlanUpdate, db: Session = Depends(get_db)):
    _get_project_or_404(db, project_id)
    rec = db.get(ClientPaymentsPlan, pay_id)
    if not rec or rec.project_id != project_id:
        raise HTTPException(404, "PAYMENT_PLAN_NOT_FOUND")
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(rec, k, v)
    db.commit()
    db.refresh(rec)
    return rec

@router.delete("/{project_id}/payments/plan/{pay_id}")
def delete_payment_plan(project_id: int, pay_id: int, db: Session = Depends(get_db)):
    _get_project_or_404(db, project_id)
    rec = db.get(ClientPaymentsPlan, pay_id)
    if not rec or rec.project_id != project_id:
        raise HTTPException(404, "PAYMENT_PLAN_NOT_FOUND")
    db.delete(rec)
    db.commit()
    return {"deleted": True}

@router.get("/{project_id}/payments/fact", response_model=list[PaymentFactOut])
def list_payments_fact(project_id: int, db: Session = Depends(get_db)):
    _get_project_or_404(db, project_id)
    today = date.today()
    facts = db.execute(
        select(ClientPaymentsFact).where(ClientPaymentsFact.project_id == project_id).order_by(ClientPaymentsFact.pay_date.asc(), ClientPaymentsFact.id.asc())
    ).scalars().all()
    due_plans = db.execute(
        select(ClientPaymentsPlan).where(
            ClientPaymentsPlan.project_id == project_id,
            ClientPaymentsPlan.pay_date <= today,
        ).order_by(ClientPaymentsPlan.pay_date.asc(), ClientPaymentsPlan.id.asc())
    ).scalars().all()

    rows: list[dict] = [
        {
            "id": int(f.id),
            "project_id": int(f.project_id),
            "pay_date": f.pay_date,
            "amount": float(f.amount),
            "note": f.note,
        }
        for f in facts
    ]
    rows.extend([
        {
            "id": -int(p.id),
            "project_id": int(p.project_id),
            "pay_date": p.pay_date,
            "amount": float(p.amount),
            "note": p.note,
        }
        for p in due_plans
    ])
    rows.sort(key=lambda r: (r["pay_date"], r["id"]))
    return rows

@router.post("/{project_id}/payments/fact", response_model=PaymentFactOut)
def create_payment_fact(project_id: int, payload: PaymentFactCreate, db: Session = Depends(get_db)):
    _get_project_or_404(db, project_id)
    rec = ClientPaymentsFact(
        project_id=project_id,
        pay_date=payload.pay_date,
        amount=payload.amount,
        note=payload.note,
    )
    db.add(rec)
    db.commit()
    db.refresh(rec)
    return rec

@router.patch("/{project_id}/payments/fact/{fact_id}", response_model=PaymentFactOut)
def update_payment_fact(project_id: int, fact_id: int, payload: PaymentFactUpdate, db: Session = Depends(get_db)):
    _get_project_or_404(db, project_id)
    if fact_id < 0:
        plan_id = -fact_id
        rec_plan = db.get(ClientPaymentsPlan, plan_id)
        if not rec_plan or rec_plan.project_id != project_id:
            raise HTTPException(404, "PAYMENT_FACT_NOT_FOUND")
        for k, v in payload.model_dump(exclude_unset=True).items():
            setattr(rec_plan, k, v)
        db.commit()
        db.refresh(rec_plan)
        return {
            "id": -int(rec_plan.id),
            "project_id": int(rec_plan.project_id),
            "pay_date": rec_plan.pay_date,
            "amount": float(rec_plan.amount),
            "note": rec_plan.note,
        }

    rec = db.get(ClientPaymentsFact, fact_id)
    if not rec or rec.project_id != project_id:
        raise HTTPException(404, "PAYMENT_FACT_NOT_FOUND")
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(rec, k, v)
    db.commit()
    db.refresh(rec)
    return rec

@router.delete("/{project_id}/payments/fact/{fact_id}")
def delete_payment_fact(project_id: int, fact_id: int, db: Session = Depends(get_db)):
    _get_project_or_404(db, project_id)
    if fact_id < 0:
        plan_id = -fact_id
        rec_plan = db.get(ClientPaymentsPlan, plan_id)
        if not rec_plan or rec_plan.project_id != project_id:
            raise HTTPException(404, "PAYMENT_FACT_NOT_FOUND")
        db.delete(rec_plan)
        db.commit()
        return {"deleted": True}

    rec = db.get(ClientPaymentsFact, fact_id)
    if not rec or rec.project_id != project_id:
        raise HTTPException(404, "PAYMENT_FACT_NOT_FOUND")
    db.delete(rec)
    db.commit()
    return {"deleted": True}
