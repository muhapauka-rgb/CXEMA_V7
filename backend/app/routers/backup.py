from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import Response
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from ..db import get_db
from ..settings import settings
from ..models import (
    AdjustmentType,
    AppSettings,
    BackupFrequency,
    ClientBillingAdjustment,
    ClientPaymentsFact,
    ClientPaymentsPlan,
    ExpenseGroup,
    ExpenseItem,
    GoogleSheetLink,
    ImportItemCandidate,
    ImportJob,
    ImportRowRaw,
    ItemMode,
    Project,
    UsnMode,
)

router = APIRouter(prefix="/api/backup", tags=["backup"])
_RETENTION_MONTHS = 4
_BACKUP_NAME_RE = re.compile(r"^cxema-backup-(\d{8})-(\d{6})\.zip$")


def _backend_root() -> Path:
    return Path(__file__).resolve().parents[2]


def _resolve_path(raw_path: str) -> Path:
    path = Path(raw_path)
    if path.is_absolute():
        return path
    return (_backend_root() / path).resolve()


def backup_storage_dir() -> Path:
    db_path = _resolve_path(settings.DB_PATH)
    target = db_path.parent / "backups"
    target.mkdir(parents=True, exist_ok=True)
    return target


def _safe_copy_name(copy_name: str) -> str:
    value = Path(str(copy_name or "")).name
    if not value or value != copy_name:
        raise HTTPException(status_code=422, detail="BACKUP_COPY_INVALID")
    return value


def _parse_name_timestamp(name: str) -> Optional[datetime]:
    m = _BACKUP_NAME_RE.match(name)
    if not m:
        return None
    try:
        return datetime.strptime(f"{m.group(1)}{m.group(2)}", "%Y%m%d%H%M%S")
    except Exception:
        return None


def _month_shift(dt: datetime, months_delta: int) -> datetime:
    month0 = (dt.month - 1) + months_delta
    year = dt.year + month0 // 12
    month = month0 % 12 + 1
    day = min(dt.day, [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1])
    return dt.replace(year=year, month=month, day=day)


def list_backup_copies() -> list[dict[str, Any]]:
    root = backup_storage_dir()
    out: list[dict[str, Any]] = []
    for path in root.glob("cxema-backup-*.zip"):
        try:
            stat = path.stat()
        except FileNotFoundError:
            continue
        ts = _parse_name_timestamp(path.name)
        created = ts or datetime.fromtimestamp(stat.st_mtime)
        out.append(
            {
                "name": path.name,
                "created_at": created.isoformat(),
                "size_bytes": int(stat.st_size),
            }
        )
    out.sort(key=lambda x: x["created_at"], reverse=True)
    return out


def _copy_path_or_404(copy_name: str) -> Path:
    if copy_name == "latest":
        copies = list_backup_copies()
        if not copies:
            raise HTTPException(status_code=404, detail="BACKUP_COPY_NOT_FOUND")
        name = copies[0]["name"]
    else:
        name = _safe_copy_name(copy_name)
    target = backup_storage_dir() / name
    if not target.exists() or not target.is_file():
        raise HTTPException(status_code=404, detail="BACKUP_COPY_NOT_FOUND")
    return target


def prune_backups_older_than_months(months: int = _RETENTION_MONTHS) -> int:
    cutoff = _month_shift(datetime.utcnow(), -int(max(1, months)))
    removed = 0
    for entry in list_backup_copies():
        created = _parse_datetime(entry.get("created_at")) or datetime.utcnow()
        if created < cutoff:
            try:
                (backup_storage_dir() / entry["name"]).unlink(missing_ok=True)
                removed += 1
            except Exception:
                continue
    return removed


def _as_iso(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _parse_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value))


def _parse_datetime(value: Any) -> Optional[datetime]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    text = str(value)
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    return datetime.fromisoformat(text)


def _export_payload(db: Session) -> dict[str, Any]:
    settings = db.get(AppSettings, 1)
    projects = db.execute(select(Project).order_by(Project.id)).scalars().all()
    groups = db.execute(select(ExpenseGroup).order_by(ExpenseGroup.id)).scalars().all()
    items = db.execute(select(ExpenseItem).order_by(ExpenseItem.id)).scalars().all()
    adjustments = db.execute(select(ClientBillingAdjustment).order_by(ClientBillingAdjustment.id)).scalars().all()
    plan = db.execute(select(ClientPaymentsPlan).order_by(ClientPaymentsPlan.id)).scalars().all()
    fact = db.execute(select(ClientPaymentsFact).order_by(ClientPaymentsFact.id)).scalars().all()
    links = db.execute(select(GoogleSheetLink).order_by(GoogleSheetLink.id)).scalars().all()

    return {
        "schema_version": 1,
        "exported_at": datetime.utcnow().isoformat(),
        "app_settings": None if settings is None else {
            "id": int(settings.id),
            "usn_mode": str(settings.usn_mode.value if isinstance(settings.usn_mode, UsnMode) else settings.usn_mode),
            "usn_rate_percent": float(settings.usn_rate_percent or 0.0),
            "backup_frequency": str(settings.backup_frequency.value if isinstance(settings.backup_frequency, BackupFrequency) else settings.backup_frequency or "WEEKLY"),
            "created_at": _as_iso(settings.created_at),
            "updated_at": _as_iso(settings.updated_at),
        },
        "projects": [
            {
                "id": int(p.id),
                "title": p.title,
                "client_name": p.client_name,
                "client_email": p.client_email,
                "client_phone": p.client_phone,
                "google_drive_url": p.google_drive_url,
                "google_drive_folder": p.google_drive_folder,
                "project_price_total": float(p.project_price_total or 0.0),
                "expected_from_client_total": float(p.expected_from_client_total or 0.0),
                "agency_fee_percent": float(p.agency_fee_percent or 0.0),
                "agency_fee_include_in_estimate": bool(p.agency_fee_include_in_estimate),
                "created_at": _as_iso(p.created_at),
                "updated_at": _as_iso(p.updated_at),
                "closed_at": _as_iso(p.closed_at),
            }
            for p in projects
        ],
        "expense_groups": [
            {
                "id": int(g.id),
                "project_id": int(g.project_id),
                "name": g.name,
                "sort_order": int(g.sort_order or 0),
            }
            for g in groups
        ],
        "expense_items": [
            {
                "id": int(it.id),
                "stable_item_id": it.stable_item_id,
                "project_id": int(it.project_id),
                "group_id": int(it.group_id),
                "parent_item_id": int(it.parent_item_id) if it.parent_item_id is not None else None,
                "title": it.title,
                "mode": str(it.mode.value if isinstance(it.mode, ItemMode) else it.mode),
                "qty": None if it.qty is None else float(it.qty),
                "unit_price_base": None if it.unit_price_base is None else float(it.unit_price_base),
                "base_total": float(it.base_total or 0.0),
                "extra_profit_enabled": bool(it.extra_profit_enabled),
                "extra_profit_amount": float(it.extra_profit_amount or 0.0),
                "include_in_estimate": bool(it.include_in_estimate),
                "planned_pay_date": _as_iso(it.planned_pay_date),
                "created_at": _as_iso(it.created_at),
                "updated_at": _as_iso(it.updated_at),
            }
            for it in items
        ],
        "billing_adjustments": [
            {
                "id": int(adj.id),
                "expense_item_id": int(adj.expense_item_id),
                "unit_price_full": float(adj.unit_price_full or 0.0),
                "unit_price_billable": float(adj.unit_price_billable or 0.0),
                "adjustment_type": str(adj.adjustment_type.value if isinstance(adj.adjustment_type, AdjustmentType) else adj.adjustment_type),
                "reason": adj.reason or "",
                "discount_enabled": bool(adj.discount_enabled),
                "discount_amount": float(adj.discount_amount or 0.0),
            }
            for adj in adjustments
        ],
        "payments_plan": [
            {
                "id": int(pay.id),
                "stable_pay_id": pay.stable_pay_id,
                "project_id": int(pay.project_id),
                "pay_date": _as_iso(pay.pay_date),
                "amount": float(pay.amount or 0.0),
                "note": pay.note or "",
                "created_at": _as_iso(pay.created_at),
                "updated_at": _as_iso(pay.updated_at),
            }
            for pay in plan
        ],
        "payments_fact": [
            {
                "id": int(pay.id),
                "project_id": int(pay.project_id),
                "pay_date": _as_iso(pay.pay_date),
                "amount": float(pay.amount or 0.0),
                "note": pay.note or "",
                "created_at": _as_iso(pay.created_at),
            }
            for pay in fact
        ],
        "google_sheet_links": [
            {
                "id": int(link.id),
                "project_id": int(link.project_id),
                "spreadsheet_id": link.spreadsheet_id,
                "sheet_tab_name": link.sheet_tab_name,
                "last_published_at": _as_iso(link.last_published_at),
                "last_imported_at": _as_iso(link.last_imported_at),
            }
            for link in links
        ],
    }


def _filter_payload_by_projects(payload: dict[str, Any], selected_ids: set[int]) -> dict[str, Any]:
    groups = [g for g in payload.get("expense_groups", []) if int(g.get("project_id", 0)) in selected_ids]
    group_ids = {int(g.get("id")) for g in groups}

    items = [
        it for it in payload.get("expense_items", [])
        if int(it.get("project_id", 0)) in selected_ids and int(it.get("group_id", 0)) in group_ids
    ]
    item_ids = {int(it.get("id")) for it in items}

    return {
        **payload,
        "projects": [p for p in payload.get("projects", []) if int(p.get("id", 0)) in selected_ids],
        "expense_groups": groups,
        "expense_items": items,
        "billing_adjustments": [
            adj for adj in payload.get("billing_adjustments", [])
            if int(adj.get("expense_item_id", 0)) in item_ids
        ],
        "payments_plan": [pay for pay in payload.get("payments_plan", []) if int(pay.get("project_id", 0)) in selected_ids],
        "payments_fact": [pay for pay in payload.get("payments_fact", []) if int(pay.get("project_id", 0)) in selected_ids],
        "google_sheet_links": [link for link in payload.get("google_sheet_links", []) if int(link.get("project_id", 0)) in selected_ids],
    }


def _delete_projects_with_children(db: Session, project_ids: set[int]) -> None:
    if not project_ids:
        return

    group_ids = {
        int(gid) for gid in db.execute(
            select(ExpenseGroup.id).where(ExpenseGroup.project_id.in_(project_ids))
        ).scalars().all()
    }
    item_ids = {
        int(iid) for iid in db.execute(
            select(ExpenseItem.id).where(ExpenseItem.project_id.in_(project_ids))
        ).scalars().all()
    }
    import_job_ids = {
        int(jid) for jid in db.execute(
            select(ImportJob.id).where(ImportJob.project_id.in_(project_ids))
        ).scalars().all()
    }

    if import_job_ids:
        db.execute(delete(ImportRowRaw).where(ImportRowRaw.import_job_id.in_(import_job_ids)))
        db.execute(delete(ImportItemCandidate).where(ImportItemCandidate.import_job_id.in_(import_job_ids)))
        db.execute(delete(ImportJob).where(ImportJob.id.in_(import_job_ids)))

    if item_ids:
        db.execute(delete(ClientBillingAdjustment).where(ClientBillingAdjustment.expense_item_id.in_(item_ids)))
    db.execute(delete(ExpenseItem).where(ExpenseItem.project_id.in_(project_ids)))
    if group_ids:
        db.execute(delete(ExpenseGroup).where(ExpenseGroup.id.in_(group_ids)))
    db.execute(delete(ClientPaymentsPlan).where(ClientPaymentsPlan.project_id.in_(project_ids)))
    db.execute(delete(ClientPaymentsFact).where(ClientPaymentsFact.project_id.in_(project_ids)))
    db.execute(delete(GoogleSheetLink).where(GoogleSheetLink.project_id.in_(project_ids)))
    db.execute(delete(Project).where(Project.id.in_(project_ids)))


def _insert_payload(db: Session, payload: dict[str, Any], include_global_settings: bool) -> None:
    if include_global_settings:
        settings_raw = payload.get("app_settings")
        if settings_raw:
            row = db.get(AppSettings, 1) or AppSettings(id=1)
            row.usn_mode = UsnMode(str(settings_raw.get("usn_mode", "OPERATIONAL")).upper())
            row.usn_rate_percent = float(settings_raw.get("usn_rate_percent", 6.0) or 0.0)
            row.backup_frequency = BackupFrequency(str(settings_raw.get("backup_frequency", "WEEKLY")).upper())
            row.created_at = _parse_datetime(settings_raw.get("created_at")) or datetime.utcnow()
            row.updated_at = _parse_datetime(settings_raw.get("updated_at")) or datetime.utcnow()
            db.add(row)

    for raw in payload.get("projects", []):
        db.add(Project(
            id=int(raw["id"]),
            title=str(raw.get("title") or ""),
            client_name=raw.get("client_name"),
            client_email=raw.get("client_email"),
            client_phone=raw.get("client_phone"),
            google_drive_url=raw.get("google_drive_url"),
            google_drive_folder=raw.get("google_drive_folder"),
            project_price_total=float(raw.get("project_price_total", 0.0) or 0.0),
            expected_from_client_total=float(raw.get("expected_from_client_total", 0.0) or 0.0),
            agency_fee_percent=float(raw.get("agency_fee_percent", 10.0) or 0.0),
            agency_fee_include_in_estimate=bool(raw.get("agency_fee_include_in_estimate", True)),
            created_at=_parse_datetime(raw.get("created_at")) or datetime.utcnow(),
            updated_at=_parse_datetime(raw.get("updated_at")) or datetime.utcnow(),
            closed_at=_parse_date(raw.get("closed_at")),
        ))

    for raw in payload.get("expense_groups", []):
        db.add(ExpenseGroup(
            id=int(raw["id"]),
            project_id=int(raw["project_id"]),
            name=str(raw.get("name") or ""),
            sort_order=int(raw.get("sort_order", 0) or 0),
        ))

    for raw in payload.get("expense_items", []):
        mode_raw = str(raw.get("mode", "SINGLE_TOTAL"))
        db.add(ExpenseItem(
            id=int(raw["id"]),
            stable_item_id=str(raw.get("stable_item_id") or ""),
            project_id=int(raw["project_id"]),
            group_id=int(raw["group_id"]),
            parent_item_id=(None if raw.get("parent_item_id") is None else int(raw.get("parent_item_id"))),
            title=str(raw.get("title") or ""),
            mode=ItemMode(mode_raw),
            qty=(None if raw.get("qty") is None else float(raw.get("qty"))),
            unit_price_base=(None if raw.get("unit_price_base") is None else float(raw.get("unit_price_base"))),
            base_total=float(raw.get("base_total", 0.0) or 0.0),
            extra_profit_enabled=bool(raw.get("extra_profit_enabled", False)),
            extra_profit_amount=float(raw.get("extra_profit_amount", 0.0) or 0.0),
            include_in_estimate=bool(raw.get("include_in_estimate", True)),
            planned_pay_date=_parse_date(raw.get("planned_pay_date")),
            created_at=_parse_datetime(raw.get("created_at")) or datetime.utcnow(),
            updated_at=_parse_datetime(raw.get("updated_at")) or datetime.utcnow(),
        ))

    for raw in payload.get("billing_adjustments", []):
        adj_type = str(raw.get("adjustment_type", AdjustmentType.DISCOUNT.value))
        db.add(ClientBillingAdjustment(
            id=int(raw["id"]),
            expense_item_id=int(raw["expense_item_id"]),
            unit_price_full=float(raw.get("unit_price_full", 0.0) or 0.0),
            unit_price_billable=float(raw.get("unit_price_billable", 0.0) or 0.0),
            adjustment_type=AdjustmentType(adj_type),
            reason=str(raw.get("reason") or ""),
            discount_enabled=bool(raw.get("discount_enabled", False)),
            discount_amount=float(raw.get("discount_amount", 0.0) or 0.0),
        ))

    for raw in payload.get("payments_plan", []):
        db.add(ClientPaymentsPlan(
            id=int(raw["id"]),
            stable_pay_id=str(raw.get("stable_pay_id") or ""),
            project_id=int(raw["project_id"]),
            pay_date=_parse_date(raw.get("pay_date")) or date.today(),
            amount=float(raw.get("amount", 0.0) or 0.0),
            note=str(raw.get("note") or ""),
            created_at=_parse_datetime(raw.get("created_at")) or datetime.utcnow(),
            updated_at=_parse_datetime(raw.get("updated_at")) or datetime.utcnow(),
        ))

    for raw in payload.get("payments_fact", []):
        db.add(ClientPaymentsFact(
            id=int(raw["id"]),
            project_id=int(raw["project_id"]),
            pay_date=_parse_date(raw.get("pay_date")) or date.today(),
            amount=float(raw.get("amount", 0.0) or 0.0),
            note=str(raw.get("note") or ""),
            created_at=_parse_datetime(raw.get("created_at")) or datetime.utcnow(),
        ))

    for raw in payload.get("google_sheet_links", []):
        db.add(GoogleSheetLink(
            id=int(raw["id"]),
            project_id=int(raw["project_id"]),
            spreadsheet_id=str(raw.get("spreadsheet_id") or ""),
            sheet_tab_name=str(raw.get("sheet_tab_name") or "PROJECT"),
            last_published_at=_parse_datetime(raw.get("last_published_at")),
            last_imported_at=_parse_datetime(raw.get("last_imported_at")),
        ))


def _readable_specs() -> list[tuple[str, list[tuple[str, str]], str]]:
    return [
        (
            "settings",
            [
                ("usn_mode", "УСН режим"),
                ("usn_rate_percent", "УСН ставка, %"),
                ("backup_frequency", "Частота бэкапа"),
            ],
            "Настройки",
        ),
        (
            "projects",
            [
                ("id", "ID"),
                ("title", "Название"),
                ("client_name", "Организация"),
                ("client_email", "Email"),
                ("client_phone", "Телефон"),
                ("project_price_total", "Стоимость проекта"),
                ("expected_from_client_total", "Ожидаем от клиента"),
                ("agency_fee_percent", "Агентские, %"),
                ("google_drive_url", "Google Drive URL"),
                ("google_drive_folder", "Папка Drive"),
                ("closed_at", "Дата закрытия"),
                ("created_at", "Создан"),
                ("updated_at", "Обновлен"),
            ],
            "Проекты",
        ),
        (
            "groups",
            [
                ("id", "ID"),
                ("project_id", "ID проекта"),
                ("name", "Название"),
                ("sort_order", "Порядок"),
            ],
            "Группы",
        ),
        (
            "items",
            [
                ("id", "ID"),
                ("project_id", "ID проекта"),
                ("group_id", "ID группы"),
                ("parent_item_id", "ID родителя"),
                ("stable_item_id", "Стабильный ID"),
                ("title", "Статья"),
                ("mode", "Режим"),
                ("planned_pay_date", "Дата оплаты"),
                ("qty", "Шт"),
                ("unit_price_base", "Цена за ед"),
                ("base_total", "Сумма"),
                ("extra_profit_enabled", "Доп прибыль вкл"),
                ("extra_profit_amount", "Доп прибыль"),
                ("include_in_estimate", "В смету"),
                ("created_at", "Создан"),
                ("updated_at", "Обновлен"),
            ],
            "Расходы",
        ),
        (
            "adjustments",
            [
                ("id", "ID"),
                ("expense_item_id", "ID расхода"),
                ("adjustment_type", "Тип"),
                ("unit_price_full", "Цена полная"),
                ("unit_price_billable", "Цена клиенту"),
                ("discount_enabled", "Скидка вкл"),
                ("discount_amount", "Скидка"),
                ("reason", "Причина"),
            ],
            "Скидки",
        ),
        (
            "payments_plan",
            [
                ("id", "ID"),
                ("project_id", "ID проекта"),
                ("stable_pay_id", "Стабильный ID"),
                ("pay_date", "Дата оплаты"),
                ("amount", "Сумма"),
                ("note", "Примечание"),
                ("created_at", "Создан"),
                ("updated_at", "Обновлен"),
            ],
            "Оплаты план",
        ),
        (
            "payments_fact",
            [
                ("id", "ID"),
                ("project_id", "ID проекта"),
                ("pay_date", "Дата оплаты"),
                ("amount", "Сумма"),
                ("note", "Примечание"),
                ("created_at", "Создан"),
            ],
            "Оплаты факт",
        ),
    ]


def _rows_for_spec(payload: dict[str, Any], key: str) -> list[dict[str, Any]]:
    if key == "settings":
        row = payload.get("app_settings") or {}
        return [row] if row else []
    if key == "projects":
        return payload.get("projects", [])
    if key == "groups":
        return payload.get("expense_groups", [])
    if key == "items":
        return payload.get("expense_items", [])
    if key == "adjustments":
        return payload.get("billing_adjustments", [])
    if key == "payments_plan":
        return payload.get("payments_plan", [])
    if key == "payments_fact":
        return payload.get("payments_fact", [])
    return []


def _build_readable_xlsx(payload: dict[str, Any]) -> Optional[bytes]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return None

    def add_sheet(
        wb: Any,
        title: str,
        columns: list[tuple[str, str]],
        rows: list[dict[str, Any]],
    ) -> None:
        ws = wb.create_sheet(title=title[:31])
        headers = [label for _, label in columns]
        ws.append(headers)
        for row in rows:
            values: list[Any] = []
            for key, _ in columns:
                value = row.get(key)
                values.append("" if value is None else _as_iso(value))
            ws.append(values)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        header_fill = PatternFill(fill_type="solid", start_color="E9EEF5", end_color="E9EEF5")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_idx, (key, label) in enumerate(columns, start=1):
            max_len = len(label)
            for row in rows[:5000]:
                raw = row.get(key)
                text = "" if raw is None else str(_as_iso(raw))
                if len(text) > max_len:
                    max_len = len(text)
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max(10, max_len + 2), 48)

    wb = Workbook()
    wb.remove(wb.active)

    for key, columns, title in _readable_specs():
        add_sheet(wb, title, columns, _rows_for_spec(payload, key))

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _build_csv_bytes(columns: list[tuple[str, str]], rows: list[dict[str, Any]]) -> bytes:
    stream = io.StringIO()
    writer = csv.writer(stream, delimiter=";")
    writer.writerow([label for _, label in columns])
    for row in rows:
        writer.writerow(["" if row.get(key) is None else _as_iso(row.get(key)) for key, _ in columns])
    return ("\ufeff" + stream.getvalue()).encode("utf-8")


def _build_readable_csv_bundle(payload: dict[str, Any]) -> dict[str, bytes]:
    out: dict[str, bytes] = {}
    for key, columns, _ in _readable_specs():
        rows = _rows_for_spec(payload, key)
        out[f"readable_{key}.csv"] = _build_csv_bytes(columns, rows)
    return out


def _build_backup_zip(payload: dict[str, Any]) -> bytes:
    stamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    data_bytes = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    xlsx_bytes = _build_readable_xlsx(payload)
    manifest = {
        "format": "cxema-backup-zip",
        "schema_version": int(payload.get("schema_version", 1) or 1),
        "created_at": payload.get("exported_at") or datetime.utcnow().isoformat(),
        "files": [{"name": "data.json", "type": "application/json", "size": len(data_bytes)}],
        "backup_name": f"cxema-backup-{stamp}.zip",
    }

    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("data.json", data_bytes)
        if xlsx_bytes is not None:
            manifest["files"].append(
                {
                    "name": "readable.xlsx",
                    "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "size": len(xlsx_bytes),
                }
            )
            zf.writestr("readable.xlsx", xlsx_bytes)
        else:
            csv_bundle = _build_readable_csv_bundle(payload)
            for name, content in csv_bundle.items():
                manifest["files"].append({"name": name, "type": "text/csv", "size": len(content)})
                zf.writestr(name, content)

        manifest_bytes = json.dumps(manifest, ensure_ascii=False, indent=2).encode("utf-8")
        zf.writestr("manifest.json", manifest_bytes)
    return out.getvalue()


def build_backup_archive(db: Session) -> tuple[str, bytes]:
    payload = _export_payload(db)
    stamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    filename = f"cxema-backup-{stamp}.zip"
    content = _build_backup_zip(payload)
    return filename, content


def save_backup_to_disk(db: Session) -> Path:
    filename, content = build_backup_archive(db)
    target = backup_storage_dir() / filename
    target.write_bytes(content)
    prune_backups_older_than_months(_RETENTION_MONTHS)
    return target


def _parse_backup_bytes(raw: bytes) -> dict[str, Any]:
    if raw.startswith(b"PK"):
        try:
            with zipfile.ZipFile(io.BytesIO(raw), "r") as zf:
                target_name = "data.json" if "data.json" in zf.namelist() else next(
                    (name for name in zf.namelist() if name.lower().endswith(".json")),
                    None,
                )
                if not target_name:
                    raise HTTPException(status_code=422, detail="BACKUP_ZIP_DATA_JSON_NOT_FOUND")
                payload_raw = zf.read(target_name)
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=422, detail="BACKUP_FILE_INVALID_ZIP") from exc
        raw = payload_raw

    try:
        parsed = json.loads(raw.decode("utf-8"))
    except Exception as exc:
        raise HTTPException(status_code=422, detail="BACKUP_FILE_INVALID_JSON") from exc
    if not isinstance(parsed, dict):
        raise HTTPException(status_code=422, detail="BACKUP_FILE_INVALID_FORMAT")
    return parsed


@router.get("/export")
def export_backup(db: Session = Depends(get_db)):
    target = save_backup_to_disk(db)
    filename = target.name
    content = target.read_bytes()
    settings_row = db.get(AppSettings, 1)
    if settings_row:
        settings_row.last_backup_at = datetime.utcnow()
        db.commit()
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return Response(content=content, media_type="application/zip", headers=headers)


@router.get("/copies")
def backup_copies():
    copies = list_backup_copies()
    return {
        "retention_months": _RETENTION_MONTHS,
        "copies": copies,
        "latest": copies[0] if copies else None,
    }


@router.get("/copies/{copy_name}/projects")
def backup_copy_projects(copy_name: str):
    target = _copy_path_or_404(copy_name)
    payload = _parse_backup_bytes(target.read_bytes())
    projects = payload.get("projects", [])
    out = [
        {
            "id": int(p.get("id")),
            "title": str(p.get("title") or ""),
            "organization": p.get("client_name") or "",
        }
        for p in projects
        if p.get("id") is not None
    ]
    out.sort(key=lambda x: x["title"].lower())
    return {
        "copy_name": target.name,
        "projects": out,
    }


@router.post("/restore")
def restore_from_copy(
    copy_name: str = Query(default="latest"),
    mode: str = Query(default="full", pattern="^(full|partial)$"),
    dry_run: bool = Query(default=False),
    project_ids: Optional[str] = Query(default=None, description="CSV ids for partial import"),
    db: Session = Depends(get_db),
):
    target = _copy_path_or_404(copy_name)
    payload = _parse_backup_bytes(target.read_bytes())

    selected_ids: set[int] = set()
    if mode == "partial":
        if not project_ids:
            raise HTTPException(status_code=422, detail="PROJECT_IDS_REQUIRED_FOR_PARTIAL")
        try:
            selected_ids = {int(x.strip()) for x in str(project_ids).split(",") if x.strip()}
        except Exception as exc:
            raise HTTPException(status_code=422, detail="PROJECT_IDS_INVALID") from exc
        if not selected_ids:
            raise HTTPException(status_code=422, detail="PROJECT_IDS_EMPTY")
        payload = _filter_payload_by_projects(payload, selected_ids)

    summary = {
        "copy_name": target.name,
        "mode": mode,
        "dry_run": dry_run,
        "counts": {
            "projects": len(payload.get("projects", [])),
            "groups": len(payload.get("expense_groups", [])),
            "items": len(payload.get("expense_items", [])),
            "adjustments": len(payload.get("billing_adjustments", [])),
            "payments_plan": len(payload.get("payments_plan", [])),
            "payments_fact": len(payload.get("payments_fact", [])),
            "sheet_links": len(payload.get("google_sheet_links", [])),
        },
        "project_titles": [str(p.get("title") or "") for p in payload.get("projects", [])],
        "schema_version": int(payload.get("schema_version", 0) or 0),
    }

    if dry_run:
        return summary

    project_ids_in_payload = {int(p.get("id")) for p in payload.get("projects", []) if p.get("id") is not None}
    if mode == "full":
        _delete_projects_with_children(db, project_ids_in_payload)
        db.execute(delete(AppSettings).where(AppSettings.id == 1))
    else:
        _delete_projects_with_children(db, project_ids_in_payload)

    _insert_payload(db, payload, include_global_settings=(mode == "full"))
    db.commit()
    return {**summary, "imported": True}


@router.post("/import")
async def import_backup(
    file: UploadFile = File(...),
    mode: str = Query(default="full", pattern="^(full|partial)$"),
    dry_run: bool = Query(default=False),
    project_ids: Optional[str] = Query(default=None, description="CSV ids for partial import"),
    db: Session = Depends(get_db),
):
    raw = await file.read()
    payload = _parse_backup_bytes(raw)

    selected_ids: set[int] = set()
    if mode == "partial":
        if not project_ids:
            raise HTTPException(status_code=422, detail="PROJECT_IDS_REQUIRED_FOR_PARTIAL")
        try:
            selected_ids = {int(x.strip()) for x in str(project_ids).split(",") if x.strip()}
        except Exception as exc:
            raise HTTPException(status_code=422, detail="PROJECT_IDS_INVALID") from exc
        if not selected_ids:
            raise HTTPException(status_code=422, detail="PROJECT_IDS_EMPTY")
        payload = _filter_payload_by_projects(payload, selected_ids)

    summary = {
        "mode": mode,
        "dry_run": dry_run,
        "counts": {
            "projects": len(payload.get("projects", [])),
            "groups": len(payload.get("expense_groups", [])),
            "items": len(payload.get("expense_items", [])),
            "adjustments": len(payload.get("billing_adjustments", [])),
            "payments_plan": len(payload.get("payments_plan", [])),
            "payments_fact": len(payload.get("payments_fact", [])),
            "sheet_links": len(payload.get("google_sheet_links", [])),
        },
        "project_titles": [str(p.get("title") or "") for p in payload.get("projects", [])],
        "schema_version": int(payload.get("schema_version", 0) or 0),
    }

    if dry_run:
        return summary

    project_ids_in_payload = {int(p.get("id")) for p in payload.get("projects", []) if p.get("id") is not None}
    if mode == "full":
        _delete_projects_with_children(db, project_ids_in_payload)
        db.execute(delete(AppSettings).where(AppSettings.id == 1))
    else:
        _delete_projects_with_children(db, project_ids_in_payload)

    _insert_payload(db, payload, include_global_settings=(mode == "full"))
    db.commit()
    return {**summary, "imported": True}
